"""Text extraction from various file formats.

Supports:
- Plain text files (.txt, .md)
- CSV files (.csv)
- JSON files (.json)
- PDF files (.pdf)
- Excel files (.xlsx, .xls)
"""

from pathlib import Path

# MIME type mapping based on file extensions
EXTENSION_TO_MIME: dict[str, str] = {
    ".txt": "text/plain",
    ".md": "text/plain",
    ".csv": "text/csv",
    ".json": "application/json",
    ".pdf": "application/pdf",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls": "application/vnd.ms-excel",
}


def detect_mime_type(file_path: Path) -> str:
    """Detect MIME type of a file based on its extension.

    Args:
        file_path: Path to the file

    Returns:
        MIME type string (e.g., "text/plain", "application/pdf")
    """
    suffix = file_path.suffix.lower()
    return EXTENSION_TO_MIME.get(suffix, "application/octet-stream")


def detect_mime_type_from_filename(filename: str) -> str:
    """Detect MIME type from a filename.

    Args:
        filename: Filename with extension

    Returns:
        MIME type string
    """
    return detect_mime_type(Path(filename))


def extract_text(file_path: Path, mime_type: str | None = None) -> str | None:
    """Extract text content from a file based on its type.

    Args:
        file_path: Path to the file
        mime_type: Optional MIME type (auto-detected if not provided)

    Returns:
        Extracted text content, or None if extraction failed
    """
    if mime_type is None:
        mime_type = detect_mime_type(file_path)

    # Plain text files
    if mime_type in ("text/plain", "text/csv"):
        return _extract_text_file(file_path)

    # PDF files
    if mime_type == "application/pdf":
        return _extract_pdf_text(file_path)

    # Excel files
    if mime_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        return _extract_xlsx_text(file_path)

    # JSON files (treat as text)
    if mime_type == "application/json":
        return _extract_text_file(file_path)

    # Unknown type - try reading as text
    try:
        return _extract_text_file(file_path)
    except UnicodeDecodeError:
        return None


def extract_text_from_bytes(
    content: bytes, mime_type: str | None = None, filename: str | None = None
) -> str | None:
    """Extract text content from bytes based on MIME type.

    Args:
        content: File content as bytes
        mime_type: Optional MIME type (detected from filename if not provided)
        filename: Optional filename for MIME type detection

    Returns:
        Extracted text content, or None if extraction failed
    """
    if mime_type is None and filename:
        mime_type = detect_mime_type_from_filename(filename)
    elif mime_type is None:
        mime_type = "application/octet-stream"

    # Plain text files
    if mime_type in ("text/plain", "text/csv", "application/json"):
        try:
            return content.decode("utf-8")
        except UnicodeDecodeError:
            return content.decode("latin-1")

    # PDF files
    if mime_type == "application/pdf":
        return _extract_pdf_text_from_bytes(content)

    # Excel files
    if mime_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        return _extract_xlsx_text_from_bytes(content)

    # Unknown type - try reading as text
    try:
        return content.decode("utf-8")
    except UnicodeDecodeError:
        return None


def _extract_text_file(file_path: Path) -> str:
    """Extract text from a plain text file.

    Args:
        file_path: Path to the file

    Returns:
        File content as string
    """
    try:
        return file_path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return file_path.read_text(encoding="latin-1")


def _extract_pdf_text(file_path: Path) -> str:
    """Extract text from a PDF file.

    Args:
        file_path: Path to the PDF file

    Returns:
        Extracted text from all pages
    """
    from pypdf import PdfReader

    reader = PdfReader(file_path)
    text_parts = []

    for page in reader.pages:
        page_text = page.extract_text()
        if page_text:
            text_parts.append(page_text)

    return "\n\n".join(text_parts)


def _extract_pdf_text_from_bytes(content: bytes) -> str:
    """Extract text from PDF content in memory.

    Args:
        content: PDF file content as bytes

    Returns:
        Extracted text from all pages
    """
    import io

    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(content))
    text_parts = []

    for page in reader.pages:
        page_text = page.extract_text()
        if page_text:
            text_parts.append(page_text)

    return "\n\n".join(text_parts)


def _extract_xlsx_text(file_path: Path) -> str:
    """Extract text from an Excel file.

    Args:
        file_path: Path to the Excel file

    Returns:
        CSV-like text representation of all sheets
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    text_parts = []

    for sheet in wb.worksheets:
        sheet_lines = []
        for row in sheet.iter_rows(values_only=True):
            # Convert each cell to string, handle None
            row_values = [str(cell) if cell is not None else "" for cell in row]
            row_text = ", ".join(row_values)
            if row_text.strip(", "):  # Skip empty rows
                sheet_lines.append(row_text)

        if sheet_lines:
            # Add sheet name as header
            text_parts.append(f"[{sheet.title}]")
            text_parts.extend(sheet_lines)
            text_parts.append("")  # Empty line between sheets

    wb.close()
    return "\n".join(text_parts)


def _extract_xlsx_text_from_bytes(content: bytes) -> str:
    """Extract text from Excel content in memory.

    Args:
        content: Excel file content as bytes

    Returns:
        CSV-like text representation of all sheets
    """
    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    text_parts = []

    for sheet in wb.worksheets:
        sheet_lines = []
        for row in sheet.iter_rows(values_only=True):
            row_values = [str(cell) if cell is not None else "" for cell in row]
            row_text = ", ".join(row_values)
            if row_text.strip(", "):
                sheet_lines.append(row_text)

        if sheet_lines:
            text_parts.append(f"[{sheet.title}]")
            text_parts.extend(sheet_lines)
            text_parts.append("")

    wb.close()
    return "\n".join(text_parts)


def get_file_size(file_path: Path) -> int:
    """Get file size in bytes.

    Args:
        file_path: Path to the file

    Returns:
        File size in bytes
    """
    return file_path.stat().st_size
